"""Excel export helper using openpyxl (no pandas dependency)."""
import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def _write_sheet(ws, title: str, columns: list[str], rows: list[dict]) -> None:
    ws.title = title

    # Header
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))

    # Auto-width
    for col_idx, col_name in enumerate(columns, start=1):
        max_len = max(
            len(str(col_name)),
            *(len(str(row.get(col_name, "") or "")) for row in rows),
        ) if rows else len(col_name)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)


def gerar_excel_relatorio1(relatorio: dict) -> bytes:
    wb = Workbook()

    # Sheet 1 — Por formulário
    ws1 = wb.active
    _write_sheet(ws1, "Por Formulário", ["form_name", "total"], relatorio["por_formulario"])

    # Sheet 2 — Por usuário
    ws2 = wb.create_sheet("Por Usuário")
    forms = relatorio.get("forms_names", [])
    cols_usuario = ["user_name", "total"] + forms
    _write_sheet(ws2, "Por Usuário", cols_usuario, relatorio["por_usuario"])

    # Sheet 3 — Cruzamento
    ws3 = wb.create_sheet("Cruzamento")
    cruzamento = relatorio["cruzamento"]
    cols_cruzamento = ["Técnico"] + cruzamento["forms"]
    rows_cruzamento = [
        {"Técnico": r["user_name"], **{f: v for f, v in zip(cruzamento["forms"], r["values"])}}
        for r in cruzamento["rows"]
    ]
    _write_sheet(ws3, "Cruzamento", cols_cruzamento, rows_cruzamento)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def gerar_excel_relatorio2(relatorio: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    _write_sheet(ws, "Atividades por Usuário", relatorio["colunas"], relatorio["linhas"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
